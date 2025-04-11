import random
import tkinter as tk
from tkinter import messagebox, filedialog
from reportlab.lib.pagesizes import letter
from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, PageBreak
from reportlab.pdfbase.ttfonts import TTFont
from reportlab.lib import colors
from reportlab.platypus import Image
from reportlab.platypus import Table, TableStyle
from PIL import Image as PILImage
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.pdfbase import pdfmetrics
from openpyxl import load_workbook
from io import BytesIO

tk._default_root = None
kruti_dev_font_path = r'C:\ISHITA\KrutiDev-010.TTF'
pdfmetrics.registerFont(TTFont('KrutiDev010', kruti_dev_font_path))
excel_file_path = r'C:\ISHITA\Questi.xlsx'
wb = load_workbook(excel_file_path)
sheet = wb.active
root = tk.Tk()
root.title("Question Paper and Answer Key Generator")
window_width = 850
window_height = 600
screen_width = root.winfo_screenwidth()
screen_height = root.winfo_screenheight()
x = (screen_width // 2) - (window_width // 2)
y = (screen_height // 2) - (window_height // 2)
root.geometry(f'{window_width}x{window_height}+{x}+{y}')
root.resizable(False, False)

questions_by_type = {
    'Aptitude': [],
    'Reasoning': [],
    'Quantitative': [],
    'Verbal': [],
    'nonverbal': [],
    'SD': [],
    'MA': [],
    'RS': [],
    'Qualitative': [],
    'Coginitive': []
    # Add more categories as needed
}

for row in sheet.iter_rows(min_row=2, values_only=True):
    print("Row length:", len(row))  # Add this line for debugging
    
    if len(row) == 16:  # Adjust the condition based on your actual row structure
        sno, english_question, hindi_question, options, hindi_option, correct_option, question_type, language_indicator = row

        if language_indicator == 'N':
            options = list(options[:4])

            if isinstance(correct_option, (list, tuple)):
                correct_option = list(correct_option[:4])
            else:
                correct_option = [correct_option]

            if question_type is not None:
                questions_by_type[question_type].append((
                    english_question, hindi_question, options,hindi_option, correct_option, question_type, language_indicator
                ))
    else:
        print(f"Skipping row with unexpected length: {row}")

# The second loop should be indented here
for row in sheet.iter_rows(min_row=2, values_only=True):
    sno = row[0]
    english_question = row[1]
    hindi_question = row[6]
    options = row[2:6]
    hindi_option=row[7:11]
    correct_option = row[11:15]
    question_type = row[15]
    language_indicator = row[16]
    if question_type is not None:
        questions_by_type[question_type].append((english_question, hindi_question, options,hindi_option ,correct_option, question_type, language_indicator))

styles = getSampleStyleSheet()
styles.add(ParagraphStyle(name='HindiStyle', parent=styles['Normal'],
                          fontName='KrutiDev010', fontSize=12, leading=14))

for question_type, type_questions in questions_by_type.items():
    random.shuffle(type_questions)

# Create three sets of questions with unique orderings and shuffle within each
questions_sets = [[] for _ in range(3)]
for question_type, type_questions in questions_by_type.items():
    random.shuffle(type_questions)
    selected_questions = type_questions[:3]
    for i, set_questions in enumerate(questions_sets):
        set_questions.extend(selected_questions)

correct_options_sets = [[] for _ in range(3)]
for set_index, set_questions in enumerate(questions_sets):
    random.shuffle(set_questions)
    correct_options = [correct_option for (_, _, _, _, correct_option, _, _) in set_questions]
    correct_options = [item[i] for item in correct_options for i in range(len(item))]
    correct_options_sets[set_index].extend(correct_options)


def calculate_image_dimensions(image, page_size, frame_width, frame_height):
    max_width, max_height = page_size
    image_width, image_height = image.size
    width_scale = frame_width / image_width
    height_scale = frame_height / image_height
    scale = min(width_scale, height_scale)
    new_width = int(image_width * scale)
    new_height = int(image_height * scale)
    return new_width, new_height


def generate_instruction_content(set_index=None):
    instruction_content = []
    custom_instruction_file = filedialog.askopenfilename(
        title="Select Instruction JPG Image"
    )
    print("Selected file path:", custom_instruction_file)

    if custom_instruction_file:
        image = PILImage.open(custom_instruction_file)
        page_size = letter
        frame_width = 456
        frame_height = 636
        new_width, new_height = calculate_image_dimensions(
            image, page_size, frame_width, frame_height
        )
        image = image.resize((new_width, new_height))
        image_stream = BytesIO()
        image.save(image_stream, format='JPEG')
        custom_instruction_page = Image(
            image_stream, width=new_width, height=new_height
        )
        instruction_content.append(custom_instruction_page)

    return instruction_content


custom_paragraph_style = ParagraphStyle(
    name='MyStyle',
    parent=styles['Normal'],
    alignment=1,
    leading=14,
    spaceAfter=12,
    keepWithNext=True
)



def generate_question_pdf_for_set(set_index, instruction_images):
    try:
        # Specify the directory for question papers
        output_directory = fr'C:\ISHITA\Output\Output_Set_{set_index + 1}.pdf'
        pdf_filename = f'{output_directory}Output_Set_{set_index + 1}.pdf'
        doc = SimpleDocTemplate(pdf_filename, pagesize=letter)
        pdf_content = []

        for image in instruction_images:
            pdf_content.append(image)
            pdf_content.append(PageBreak())

        custom_paragraph = Paragraph(
            "DRDO QUESTION PAPER.", custom_paragraph_style
        )
        pdf_content.append(custom_paragraph)

        for i, question_data in enumerate(questions_sets[set_index], start=1):
           if len(question_data) >= 4:
               english_question, hindi_question, options, hindi_option, correct_option, _, language_indicator = question_data[:7]
               correct_option = list(correct_option[:7])  # Convert tuple to list
               # Check the value of language_indicator
               if language_indicator == 'N':
                   # Include both English and Hindi questions along with their respective options
                   pdf_content.append(Paragraph(f'{i}. {english_question}', styles['Normal']))
                   for idx, option in enumerate(options, start=1):
                       pdf_content.append(Paragraph(f'{chr(96 + idx)}. {option}', styles['Normal']))

                   # Add space between Hindi and English questions
                   pdf_content.append(Spacer(1, 6))

                   if hindi_question is not None and not isinstance(hindi_question, (int, float)):
                       print("Adding Hindi Question:", hindi_question)
                       pdf_content.append(Paragraph(hindi_question, styles['HindiStyle']))
                       option_label = ['1', '2', '3', '4']
                       for idx, option in enumerate(hindi_option, start=1):
                           option_text = f'{option_label[idx-1]}. {option}'
                           pdf_content.append(Paragraph(option_text, styles['HindiStyle']))
            
               else:
                   # Include only English questions and options
                   pdf_content.append(Paragraph(f'{i}. {english_question}', styles['Normal']))
                   for idx, option in enumerate(options, start=1):
                       pdf_content.append(Paragraph(f'{chr(96 + idx)}. {option}', styles['Normal']))

               pdf_content.append(Spacer(1, 12))

        doc.build(pdf_content)
        print(f'Questions PDF for Set {set_index + 1} generated: {pdf_filename}')
        messagebox.showinfo(
            "Success", f"Question PDF for Set {set_index + 1} generated successfully."
        )

    except Exception as e:
        messagebox.showerror("Error", f"An error occurred: {str(e)}")


def generate_answer_key_pdf_for_set(set_index):
    try:
        # Specify the directory for answer keys
        answer_key_directory = fr'C:\ISHITA\Output\AnswerKey_Set_{set_index + 1}.pdf'
        pdf_filename = f'{answer_key_directory}AnswerKey_Set_{set_index + 1}.pdf'
        doc = SimpleDocTemplate(pdf_filename, pagesize=letter)
        pdf_content = []
        data = [['Q_no'] + [f'K_{chr(65 + i)}' for i in range(4)]]

        for sno, question_data in enumerate(questions_sets[set_index], start=1):
            if len(question_data) >= 4:
                english_question, hindi_question,hindi_option, options, correct_option, _, _ = question_data[:7]
                correct_option = list(correct_option[:4])  # Convert tuple to list
                data.append([sno] + correct_option)

        table = Table(data, colWidths=[doc.width / 5.0] * 5)
        table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
            ('GRID', (0, 0), (-1, -1), 1, colors.black)
        ]))

        pdf_content.append(table)
        doc.build(pdf_content)
        print(f'Answer Key PDF for Set {set_index + 1} generated: {pdf_filename}')

    except Exception as e:
        messagebox.showinfo("Success", f"Answer Key PDF for Set {set_index + 1} generated successfully.")


# Set custom padding and margin values
label_padding = 5
entry_padding = 5
button_padding = 10


def generate_question_and_answer_pdfs():
    instruction_content = generate_instruction_content()  # Generate the instruction image

    for set_index in range(3):
        generate_question_pdf_for_set(set_index, instruction_content)
        generate_answer_key_pdf_for_set(set_index)
    root.update()


# Function to close the application
def close_application():
    root.destroy()


def main():
    # Create a frame to contain all the widgets
    frame = tk.Frame(root)
    frame.pack(fill=tk.BOTH, expand=True)

    # Create a title label
    title_label = tk.Label(frame, text="Question Paper and Answer Key Generator", font=("Helvetica", 16))
    title_label.grid(row=0, column=0, columnspan=2, pady=5)

    # Create a frame for file selection
    file_frame = tk.Frame(frame)
    file_frame.grid(row=1, column=0, columnspan=2, pady=5)

    # Create a frame for buttons
    button_frame = tk.Frame(frame)
    button_frame.grid(row=2, column=0, columnspan=2, pady=10)

    # Create a button to generate both question papers and answer keys
    generate_button = tk.Button(button_frame, text="Generate Question Paper and Answer Key",
                                command=generate_question_and_answer_pdfs)
    generate_button.grid(row=0, column=0, padx=10)

    # Create a button to close the application
    close_button = tk.Button(button_frame, text="Close", command=close_application)
    close_button.grid(row=0, column=1, padx=10)
    root.mainloop()


if __name__ == "__main__":
    main()
    
    
    
    
import random
import tkinter as tk
from tkinter import messagebox, filedialog
from reportlab.lib.pagesizes import letter
from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, PageBreak
from reportlab.pdfbase.ttfonts import TTFont
from reportlab.lib import colors
from reportlab.platypus import Image
from reportlab.platypus import Table, TableStyle
from PIL import Image as PILImage
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.pdfbase import pdfmetrics
from openpyxl import load_workbook
from io import BytesIO


class QuestionPaperGenerator:
    def __init__(self, root):
        self.root = root
        self.root.title("Question Paper and Answer Key Generator")
        kruti_dev_font_path = r'C:\ISHITA\KrutiDev-010.TTF'
        pdfmetrics.registerFont(TTFont('KrutiDev010', kruti_dev_font_path))
        excel_file_path = r'C:\ISHITA\Questi.xlsx'
        wb = load_workbook(excel_file_path)
        sheet = wb.active
        self.window_width = 850
        self.window_height = 600
        self.screen_width = self.root.winfo_screenwidth()
        self.screen_height = self.root.winfo_screenheight()
        self.x = (self.screen_width // 2) - (self.window_width // 2)
        self.y = (self.screen_height // 2) - (self.window_height // 2)
        self.root.geometry(f'{self.window_width}x{self.window_height}+{self.x}+{self.y}')
        self.root.resizable(False, False)

    questions_by_type = {
        'Aptitude': [],
        'Reasoning': [],
        'Quantitative': [],
        'Verbal': [],
        'nonverbal': [],
        'SD': [],
        'MA': [],
        'RS': [],
        'Qualitative': [],
        'Coginitive': []
        # Add more categories as needed
    }
    
    
    for row in sheet.iter_rows(min_row=2, values_only=True):
        print("Row length:", len(row))  # Add this line for debugging
        
        if len(row) == 16:  # Adjust the condition based on your actual row structure
            sno, english_question, hindi_question, options, hindi_option, correct_option, question_type, language_indicator = row

            if language_indicator == 'N':
                options = list(options[:4])

                if isinstance(correct_option, (list, tuple)):
                    correct_option = list(correct_option[:4])
                else:
                    correct_option = [correct_option]

                if question_type is not None:
                    questions_by_type[question_type].append((
                        english_question, hindi_question, options,hindi_option, correct_option, question_type, language_indicator
                    ))
        else:
            print(f"Skipping row with unexpected length: {row}")

    # The second loop should be indented here
    for row in sheet.iter_rows(min_row=2, values_only=True):
        sno = row[0]
        english_question = row[1]
        hindi_question = row[6]
        options = row[2:6]
        hindi_option=row[7:11]
        correct_option = row[11:15]
        question_type = row[15]
        language_indicator = row[16]
        if question_type is not None:
            questions_by_type[question_type].append((english_question, hindi_question, options,hindi_option ,correct_option, question_type, language_indicator))

    styles = getSampleStyleSheet()
    styles.add(ParagraphStyle(name='HindiStyle', parent=styles['Normal'],
                            fontName='KrutiDev010', fontSize=12, leading=14))

    for question_type, type_questions in questions_by_type.items():
        random.shuffle(type_questions)

    # Create three sets of questions with unique orderings and shuffle within each
    questions_sets = [[] for _ in range(3)]
    for question_type, type_questions in questions_by_type.items():
        random.shuffle(type_questions)
        selected_questions = type_questions[:3]
        for i, set_questions in enumerate(questions_sets):
            set_questions.extend(selected_questions)

    correct_options_sets = [[] for _ in range(3)]
    for set_index, set_questions in enumerate(questions_sets):
        random.shuffle(set_questions)
        correct_options = [correct_option for (_, _, _, _, correct_option, _, _) in set_questions]
        correct_options = [item[i] for item in correct_options for i in range(len(item))]
        correct_options_sets[set_index].extend(correct_options)


    def calculate_image_dimensions(image, page_size, frame_width, frame_height):
        max_width, max_height = page_size
        image_width, image_height = image.size
        width_scale = frame_width / image_width
        height_scale = frame_height / image_height
        scale = min(width_scale, height_scale)
        new_width = int(image_width * scale)
        new_height = int(image_height * scale)
        return new_width, new_height


    def run(self):
        self.create_widgets()
        self.root.mainloop()

    def create_widgets(self):
        frame = tk.Frame(self.root)
        frame.pack(fill=tk.BOTH, expand=True)

        # Create a title label
        title_label = tk.Label(frame, text="Question Paper and Answer Key Generator", font=("Helvetica", 16))
        title_label.grid(row=0, column=0, columnspan=2, pady=5)

        # Create a frame for file selection
        file_frame = tk.Frame(frame)
        file_frame.grid(row=1, column=0, columnspan=2, pady=5)

        # Create a frame for buttons
        button_frame = tk.Frame(frame)
        button_frame.grid(row=2, column=0, columnspan=2, pady=10)

        # Create a button to generate both question papers and answer keys
        generate_button = tk.Button(button_frame, text="Generate Question Paper and Answer Key",
                                    command=self.generate_question_and_answer_pdfs)
        generate_button.grid(row=0, column=0, padx=10)

        # Create a button to close the application
        close_button = tk.Button(button_frame, text="Close", command=self.close_application)
        close_button.grid(row=0, column=1, padx=10)

def generate_instruction_content(set_index=None):
    instruction_content = []
    custom_instruction_file = filedialog.askopenfilename(
        title="Select Instruction JPG Image"
    )
    print("Selected file path:", custom_instruction_file)

    if custom_instruction_file:
        image = PILImage.open(custom_instruction_file)
        page_size = letter
        frame_width = 456
        frame_height = 636
        new_width, new_height = calculate_image_dimensions(
            image, page_size, frame_width, frame_height
        )
        image = image.resize((new_width, new_height))
        image_stream = BytesIO()
        image.save(image_stream, format='JPEG')
        custom_instruction_page = Image(
            image_stream, width=new_width, height=new_height
        )
        instruction_content.append(custom_instruction_page)

    return instruction_content


custom_paragraph_style = ParagraphStyle(
    name='MyStyle',
    parent=styles['Normal'],
    alignment=1,
    leading=14,
    spaceAfter=12,
    keepWithNext=True
)



def generate_question_pdf_for_set(set_index, instruction_images):
    try:
        # Specify the directory for question papers
        output_directory = fr'C:\ISHITA\Output\Output_Set_{set_index + 1}.pdf'
        pdf_filename = f'{output_directory}Output_Set_{set_index + 1}.pdf'
        doc = SimpleDocTemplate(pdf_filename, pagesize=letter)
        pdf_content = []

        for image in instruction_images:
            pdf_content.append(image)
            pdf_content.append(PageBreak())

        custom_paragraph = Paragraph(
            "DRDO QUESTION PAPER.", custom_paragraph_style
        )
        pdf_content.append(custom_paragraph)

        for i, question_data in enumerate(questions_sets[set_index], start=1):
           if len(question_data) >= 4:
               english_question, hindi_question, options, hindi_option, correct_option, _, language_indicator = question_data[:7]
               correct_option = list(correct_option[:7])  # Convert tuple to list
               # Check the value of language_indicator
               if language_indicator == 'N':
                   # Include both English and Hindi questions along with their respective options
                   pdf_content.append(Paragraph(f'{i}. {english_question}', styles['Normal']))
                   for idx, option in enumerate(options, start=1):
                       pdf_content.append(Paragraph(f'{chr(96 + idx)}. {option}', styles['Normal']))

                   # Add space between Hindi and English questions
                   pdf_content.append(Spacer(1, 6))

                   if hindi_question is not None and not isinstance(hindi_question, (int, float)):
                       print("Adding Hindi Question:", hindi_question)
                       pdf_content.append(Paragraph(hindi_question, styles['HindiStyle']))
                       option_label = ['1', '2', '3', '4']
                       for idx, option in enumerate(hindi_option, start=1):
                           option_text = f'{option_label[idx-1]}. {option}'
                           pdf_content.append(Paragraph(option_text, styles['HindiStyle']))
            
               else:
                   # Include only English questions and options
                   pdf_content.append(Paragraph(f'{i}. {english_question}', styles['Normal']))
                   for idx, option in enumerate(options, start=1):
                       pdf_content.append(Paragraph(f'{chr(96 + idx)}. {option}', styles['Normal']))

               pdf_content.append(Spacer(1, 12))

        doc.build(pdf_content)
        print(f'Questions PDF for Set {set_index + 1} generated: {pdf_filename}')
        messagebox.showinfo(
            "Success", f"Question PDF for Set {set_index + 1} generated successfully."
        )

    except Exception as e:
        messagebox.showerror("Error", f"An error occurred: {str(e)}")


def generate_answer_key_pdf_for_set(set_index):
    try:
        # Specify the directory for answer keys
        answer_key_directory = fr'C:\ISHITA\Output\AnswerKey_Set_{set_index + 1}.pdf'
        pdf_filename = f'{answer_key_directory}AnswerKey_Set_{set_index + 1}.pdf'
        doc = SimpleDocTemplate(pdf_filename, pagesize=letter)
        pdf_content = []
        data = [['Q_no'] + [f'K_{chr(65 + i)}' for i in range(4)]]

        for sno, question_data in enumerate(questions_sets[set_index], start=1):
            if len(question_data) >= 4:
                english_question, hindi_question,hindi_option, options, correct_option, _, _ = question_data[:7]
                correct_option = list(correct_option[:4])  # Convert tuple to list
                data.append([sno] + correct_option)

        table = Table(data, colWidths=[doc.width / 5.0] * 5)
        table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
            ('GRID', (0, 0), (-1, -1), 1, colors.black)
        ]))

        pdf_content.append(table)
        doc.build(pdf_content)
        print(f'Answer Key PDF for Set {set_index + 1} generated: {pdf_filename}')

    except Exception as e:
        messagebox.showinfo("Success", f"Answer Key PDF for Set {set_index + 1} generated successfully.")


# Set custom padding and margin values
label_padding = 5
entry_padding = 5
button_padding = 10


def generate_question_and_answer_pdfs():
    instruction_content = generate_instruction_content()  # Generate the instruction image

    for set_index in range(3):
        generate_question_pdf_for_set(set_index, instruction_content)
        generate_answer_key_pdf_for_set(set_index)
    root.update()


# Function to close the application
def close_application():
    root.destroy()




    def close_application(self):
        self.root.destroy()


def main():
    root = tk.Tk()
    question_paper_generator = QuestionPaperGenerator(root)
    question_paper_generator.run()


if __name__ == "__main__":
    main()
